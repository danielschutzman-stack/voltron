"""
territory_xls.py
TerritoryWorkbook — create and progressively update a territory PG plan .xlsx file.

v2 changes: No file reads or writes. save() returns bytes instead of writing
to disk. load_existing() accepts bytes instead of reading from a path.
Caller is responsible for storing and delivering the workbook bytes.

Usage:
    from territory_xls import TerritoryWorkbook

    wb = TerritoryWorkbook()

    # New session — create from scratch:
    wb.create_skeleton(accounts)

    # Resuming from previously saved bytes:
    if not wb.load_existing(saved_bytes):
        wb.create_skeleton(accounts)

    # Progressive updates — call wb.save() after every update:
    wb.update_web_research(account_name, data)
    xlsx_bytes = wb.save()   # returns bytes — deliver via platform mechanism

    wb.update_sfdc(account_name, data)
    xlsx_bytes = wb.save()

    # Mark complete:
    wb.update_pg_complete(account_name, pg_report_url="...", onepager_url="...")
    xlsx_bytes = wb.save()
"""

import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Brand colors
# ---------------------------------------------------------------------------

NAVY       = "1D2D50"
BLUE       = "2E5CE5"
LIGHT_BLUE = "EBF2FF"
WHITE      = "FFFFFF"


# ---------------------------------------------------------------------------
# Summary sheet column definitions
# ---------------------------------------------------------------------------

SUMMARY_COLUMNS = [
    "Account Name", "Segment", "Owner", "Vertical",
    "ICP Grade (TSA)", "ICP Grade (TSE)", "6Sense Intent",
    "Last Activity Date", "Days Since Touch", "Deal Stage",
    "Open Opp", "Case Studies", "PG Report", "One-Pager",
    "PG Plan Status", "Notes",
]

_6SENSE_THRESHOLDS = [
    (90, "A+"),
    (75, "A"),
    (60, "B"),
    (0,  "C"),
]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _score_to_grade(score) -> str:
    try:
        n = float(score)
        for threshold, grade in _6SENSE_THRESHOLDS:
            if n >= threshold:
                return grade
        return "C"
    except (TypeError, ValueError):
        return str(score) if score else ""


def _safe_join(items, key: str, limit: int = 3) -> str:
    parts = []
    for item in items[:limit]:
        if isinstance(item, dict):
            val = item.get(key, "")
            if val:
                parts.append(str(val))
        else:
            parts.append(str(item))
    return ", ".join(parts)


def _safe_text(items, key: str, limit: int = 2, sep: str = "; ") -> str:
    parts = []
    for item in items[:limit]:
        if isinstance(item, dict):
            val = item.get(key, "")
            if val:
                parts.append(str(val))
        else:
            parts.append(str(item))
    return sep.join(parts)


def _truncate(text: str, max_len: int = 497) -> str:
    if not text:
        return ""
    return text[:max_len] + "..." if len(text) > max_len else text


# ---------------------------------------------------------------------------
# TerritoryWorkbook
# ---------------------------------------------------------------------------

class TerritoryWorkbook:

    def __init__(self, filepath: str = "territory_pg_plan.xlsx"):
        """
        Parameters
        ----------
        filepath : Suggested filename for the workbook when delivered.
                   Not used for file I/O — kept for naming reference only.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required. Install with: pip install openpyxl"
            )
        self.filepath         = filepath
        self._account_rows    = {}
        self._wb              = None
        self._summary_ws      = None
        self._tab_module_rows = {}
        self._sheet_names     = set()

    # ── Public: initialization ───────────────────────────────────────────────

    def create_skeleton(self, accounts: list):
        """
        Create a fresh workbook with a Summary sheet and one tab per account.

        Parameters
        ----------
        accounts : list of dicts, each with keys:
            account_name, segment, owner, vertical, tsa_icp, tse_icp,
            intent_score, last_activity_date, days_since_touch, deal_stage,
            open_opp, case_studies, pg_status, notes
        """
        self._wb              = openpyxl.Workbook()
        self._summary_ws      = self._wb.active
        self._summary_ws.title = "Summary"
        self._sheet_names     = {"Summary"}

        self._build_summary_header()

        for i, acct in enumerate(accounts, start=2):
            name = acct.get("account_name", f"Account {i - 1}")
            self._account_rows[name] = i
            self._write_summary_row(i, acct)
            self._create_account_tab(name, acct)

        self._style_summary_sheet()

    def load_existing(self, data: bytes = None) -> bool:
        """
        Load a workbook from bytes.

        Returns True if loaded successfully, False if no data provided.

        Usage:
            if not wb.load_existing(saved_bytes):
                wb.create_skeleton(accounts)
        """
        if not data:
            return False
        try:
            buffer            = io.BytesIO(data)
            self._wb          = openpyxl.load_workbook(buffer)
            self._summary_ws  = self._wb["Summary"]
            self._sheet_names = set(self._wb.sheetnames)

            # Rebuild account row index from column A
            for row in self._summary_ws.iter_rows(min_row=2, max_col=1):
                cell = row[0]
                if cell.value:
                    self._account_rows[cell.value] = cell.row

            # Rebuild tab module row maps
            for account_name in self._account_rows:
                safe = self._safe_sheet_name(account_name, set())
                if safe in self._wb.sheetnames:
                    ws      = self._wb[safe]
                    row_map = {}
                    for row in ws.iter_rows(min_row=5, max_col=1):
                        cell = row[0]
                        if cell.value:
                            row_map[cell.value] = cell.row
                    self._tab_module_rows[account_name] = row_map

            return True
        except Exception:
            return False

    # ── Public: progressive update methods ──────────────────────────────────

    def update_web_research(self, account_name: str, data: dict):
        """Update from web_research subagent output."""
        vertical = data.get("industry", "")
        if vertical:
            self._update_summary_cell(account_name, "Vertical", vertical)

        priorities = data.get("strategic_priorities", [])[:3]
        summary    = _safe_text(priorities, "text", limit=3) or "See full report."

        self._update_tab_module(
            account_name, "Web Research", "✅ Complete", summary
        )

    def update_sfdc(self, account_name: str, data: dict):
        """Update from sfdc_stakeholder query output."""
        owner = data.get("account_owner", "")
        deal  = data.get("deal_stage", "")
        opp   = data.get("opportunity_name", "")

        if owner: self._update_summary_cell(account_name, "Owner", owner)
        if deal:  self._update_summary_cell(account_name, "Deal Stage", deal)
        if opp:   self._update_summary_cell(account_name, "Open Opp", opp)

        self._update_tab_module(
            account_name,
            "SFDC / Stakeholder Lookup",
            "✅ Complete",
            _truncate(f"Owner: {owner} | Deal: {deal} | Opp: {opp}"),
        )

    def update_6sense(self, account_name: str, data: dict):
        """
        Update from 6sense_intent query output.
        Converts raw score to grade — never writes raw number per PG Report rule #1.
        """
        raw_score = data.get("intent_score", data.get("reach_score", ""))
        grade     = _score_to_grade(raw_score)

        if grade:
            self._update_summary_cell(account_name, "6Sense Intent", grade)

        self._update_tab_module(
            account_name, "6Sense Intent", "✅ Complete", f"Intent Grade: {grade}"
        )

    def update_case_studies(self, account_name: str, data: dict):
        """Update from case_study_matcher subagent output."""
        studies = data.get("recommended_case_studies", [])
        names   = _safe_join(studies, "company", limit=3)

        self._update_summary_cell(account_name, "Case Studies", names)
        self._update_tab_module(
            account_name,
            "Case Study Matcher",
            "✅ Complete",
            _truncate(f"Top matches: {names}"),
        )

    def update_tsumble(self, account_name: str, data: dict):
        """Update from tsumble subagent output."""
        total  = data.get("total_open_roles", 0)
        trends = _safe_text(data.get("hiring_trends", []), "trend", limit=2)

        self._update_tab_module(
            account_name,
            "TSumbleV1 (Job Openings)",
            "✅ Complete",
            _truncate(f"{total} open roles. Trends: {trends}"),
        )

    def update_competitor_intel(self, account_name: str, data: dict):
        """Update from competitor_intel subagent output."""
        confirmed = data.get("tools_confirmed", [])
        tools     = _safe_join(confirmed, "tool", limit=5)

        self._update_tab_module(
            account_name,
            "Competitor Intel",
            "✅ Complete",
            _truncate(f"Confirmed tools: {tools}"),
        )

    def update_deal_story(self, account_name: str, data: dict):
        """Update from deal_stage + deal_funnel_timing query output."""
        stage   = data.get("deal_stage", "")
        summary = data.get("summary", f"Stage: {stage}")

        self._update_tab_module(
            account_name, "Deal Story", "✅ Complete", _truncate(summary)
        )

    def update_meddpicc(self, account_name: str, data: dict):
        """
        Update from meddpicc_flags + meddpicc_detail query output.
        MEDDPICC data is internal AE reference only — never shown in
        customer-facing HTML reports per PG Report rule #2.
        """
        score   = data.get("meddpicc_score", "")
        summary = data.get("summary", f"MEDDPICC Score: {score}")

        self._update_tab_module(
            account_name,
            "Sales Call Analyzer (MEDDPICC)",
            "✅ Complete",
            _truncate(summary),
        )

    def update_exec_profiles(self, account_name: str, data: dict):
        """Update from exec_profile subagent output."""
        executives = data.get("executives", [])
        names      = _safe_join(executives, "name", limit=3)

        self._update_tab_module(
            account_name,
            "Exec Profile Builder",
            "✅ Complete",
            _truncate(f"Profiles: {names}"),
        )

    def update_outreach(self, account_name: str, data: dict):
        """Update from Outreach Generator output."""
        self._update_tab_module(
            account_name,
            "Outreach Generator",
            "✅ Complete",
            "Email + LinkedIn sequences generated.",
        )

    def update_pg_complete(
        self,
        account_name:  str,
        pg_report_url: str = None,
        onepager_url:  str = None,
        notes:         str = None,
    ):
        """
        Mark a PG plan as complete and add report links to Summary sheet.

        Parameters
        ----------
        account_name   : Must match the name used in create_skeleton()
        pg_report_url  : URL or reference to the final PG report
        onepager_url   : URL or reference to the one-pager
        notes          : Optional notes to add to the Notes column
        """
        today = datetime.now().strftime("%Y-%m-%d")
        self._update_summary_cell(
            account_name, "PG Plan Status", f"✅ Completed — {today}"
        )

        row = self._account_rows.get(account_name)
        if not row or not self._summary_ws:
            return

        pg_col       = SUMMARY_COLUMNS.index("PG Report") + 1
        onepager_col = SUMMARY_COLUMNS.index("One-Pager") + 1
        link_font    = Font(color=BLUE, underline="single", name="Calibri", size=10)

        if pg_report_url:
            cell           = self._summary_ws.cell(row=row, column=pg_col)
            cell.value     = "📄 PG Report"
            cell.hyperlink = pg_report_url
            cell.font      = link_font

        if onepager_url:
            cell           = self._summary_ws.cell(row=row, column=onepager_col)
            cell.value     = "📋 One-Pager"
            cell.hyperlink = onepager_url
            cell.font      = link_font

        if notes:
            self._update_summary_cell(account_name, "Notes", notes)

    def save(self) -> bytes:
        """
        Serialize workbook to bytes — no file writes.
        Caller is responsible for delivering via platform mechanism.

        Returns
        -------
        bytes : Complete Excel file contents ready to save or deliver.

        Raises RuntimeError if workbook not initialized.
        """
        if self._wb is None:
            raise RuntimeError(
                "Workbook not initialized. "
                "Call create_skeleton() or load_existing() first."
            )
        buffer = io.BytesIO()
        self._wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def get_account_list(self) -> list:
        """Return list of account names in the workbook."""
        return list(self._account_rows.keys())

    # ── Internal: sheet name deduplication ──────────────────────────────────

    def _safe_sheet_name(self, account_name: str, existing: set) -> str:
        base = (
            account_name[:28]
            .replace("/", "-")
            .replace("[", "")
            .replace("]", "")
            .replace(":", "-")
            .replace("*", "")
            .replace("?", "")
            .replace("\\", "-")
        )
        name    = base
        counter = 2
        check   = existing | self._sheet_names
        while name in check:
            suffix = f" ({counter})"
            name   = base[:31 - len(suffix)] + suffix
            counter += 1
        return name

    # ── Internal: summary sheet ──────────────────────────────────────────────

    def _build_summary_header(self):
        ws          = self._summary_ws
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor=NAVY)
        header_font = Font(bold=True, color=WHITE, name="Calibri", size=11)

        for col_idx, col_name in enumerate(SUMMARY_COLUMNS, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.row_dimensions[1].height = 30

    def _write_summary_row(self, row: int, acct: dict):
        ws         = self._summary_ws
        fill_color = LIGHT_BLUE if (row % 2 == 0) else WHITE
        row_fill   = PatternFill("solid", fgColor=fill_color)
        base_font  = Font(name="Calibri", size=10)

        tsa_grade    = _score_to_grade(acct.get("tsa_icp", ""))
        tse_grade    = _score_to_grade(acct.get("tse_icp", ""))
        intent_grade = _score_to_grade(acct.get("intent_score", ""))

        values = {
            "Account Name":       acct.get("account_name", ""),
            "Segment":            acct.get("segment", ""),
            "Owner":              acct.get("owner", ""),
            "Vertical":           acct.get("vertical", ""),
            "ICP Grade (TSA)":    tsa_grade,
            "ICP Grade (TSE)":    tse_grade,
            "6Sense Intent":      intent_grade,
            "Last Activity Date": acct.get("last_activity_date", ""),
            "Days Since Touch":   acct.get("days_since_touch", ""),
            "Deal Stage":         acct.get("deal_stage", ""),
            "Open Opp":           acct.get("open_opp", ""),
            "Case Studies":       acct.get("case_studies", ""),
            "PG Report":          "—",
            "One-Pager":          "—",
            "PG Plan Status":     acct.get("pg_status", "⬜ Not Started"),
            "Notes":              acct.get("notes", ""),
        }

        for col_idx, col_name in enumerate(SUMMARY_COLUMNS, start=1):
            cell           = ws.cell(row=row, column=col_idx, value=values.get(col_name, ""))
            cell.fill      = row_fill
            cell.font      = base_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _style_summary_sheet(self):
        ws = self._summary_ws
        col_widths = {
            "Account Name":       28,
            "Segment":            12,
            "Owner":              18,
            "Vertical":           18,
            "ICP Grade (TSA)":    14,
            "ICP Grade (TSE)":    14,
            "6Sense Intent":      14,
            "Last Activity Date": 18,
            "Days Since Touch":   16,
            "Deal Stage":         14,
            "Open Opp":           20,
            "Case Studies":       20,
            "PG Report":          14,
            "One-Pager":          14,
            "PG Plan Status":     22,
            "Notes":              30,
        }
        for col_idx, col_name in enumerate(SUMMARY_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                col_widths.get(col_name, 15)
            )

    def _update_summary_cell(self, account_name: str, column_name: str, value):
        row = self._account_rows.get(account_name)
        if row is None:
            return
        try:
            col_idx = SUMMARY_COLUMNS.index(column_name) + 1
        except ValueError:
            return
        self._summary_ws.cell(row=row, column=col_idx, value=value)

    # ── Internal: account tabs ───────────────────────────────────────────────

    def _create_account_tab(self, account_name: str, acct: dict):
        safe_name = self._safe_sheet_name(account_name, set())
        self._sheet_names.add(safe_name)

        ws          = self._wb.create_sheet(title=safe_name)
        header_fill = PatternFill("solid", fgColor=NAVY)
        header_font = Font(bold=True, color=WHITE, name="Calibri", size=12)

        ws["A1"]      = account_name
        ws["A1"].font = Font(bold=True, name="Calibri", size=14, color=NAVY)
        ws["A2"]      = f"PG Plan | Generated: {datetime.now().strftime('%Y-%m-%d')}"
        ws["A2"].font = Font(italic=True, name="Calibri", size=10, color="666666")

        for cell_addr, val in [
            ("A4", "MODULE"),
            ("B4", "STATUS"),
            ("C4", "SUMMARY"),
            ("D4", "LAST UPDATED"),
        ]:
            c           = ws[cell_addr]
            c.value     = val
            c.fill      = header_fill
            c.font      = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        modules = [
            "Web Research",
            "TSumbleV1 (Job Openings)",
            "SFDC / Stakeholder Lookup",
            "Deal Story",
            "Sales Call Analyzer (MEDDPICC)",
            "Competitor Intel",
            "6Sense Intent",
            "Case Study Matcher",
            "Exec Profile Builder",
            "Outreach Generator",
        ]

        for row_offset, module in enumerate(modules, start=5):
            fill_color = LIGHT_BLUE if (row_offset % 2 == 0) else WHITE
            for col in range(1, 5):
                ws.cell(row=row_offset, column=col).fill = PatternFill(
                    "solid", fgColor=fill_color
                )
            ws.cell(row=row_offset, column=1, value=module)
            ws.cell(row=row_offset, column=2, value="⬜ Pending")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 18
        ws.freeze_panes = "A5"

        self._tab_module_rows[account_name] = {
            modules[i]: 5 + i for i in range(len(modules))
        }

    def _update_tab_module(
        self,
        account_name: str,
        module_name:  str,
        status:       str,
        summary:      str,
    ):
        row_map = self._tab_module_rows.get(account_name, {})
        row     = row_map.get(module_name)
        if row is None:
            return

        safe_name = self._safe_sheet_name(account_name, set())
        if safe_name not in self._wb.sheetnames:
            return

        ws = self._wb[safe_name]
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=_truncate(summary) if summary else "")
        ws.cell(row=row, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
